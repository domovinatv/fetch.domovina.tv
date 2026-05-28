#!/usr/bin/env python3
"""
loudness_report_xlsx.py

Pretvara per-epizoda mjerenja iz analyze_loudness.js (loudness_data.json) u
pregledni .xlsx s obojanim ćelijama. Bojanje po udaljenosti integrated LUFS od
cilja (default -16): zeleno ≤1 LU, žuto ≤2, narančasto ≤3, crveno >3.

NAPOMENA: ebur128 mjeri 16kHz mono .wav → ~+3 LU offset vs full-band stereo
izvora. RASPON između epizoda je valjan, apsolut nije (re-mjeri se u trenutku
two-pass loudnorm normalizacije). Bojanje vs -16 služi za vizualni pregled
rasipanja, ne kao konačni verdikt.

Usage:
  python3 loudness_report_xlsx.py
  python3 loudness_report_xlsx.py --data loudness_data.json --out loudness_report.xlsx --target -16
"""

import argparse
import json
import os
import re
import statistics
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

p = argparse.ArgumentParser()
p.add_argument("--data", default="loudness_data.json")
p.add_argument("--out", default="loudness_report.xlsx")
p.add_argument("--target", type=float, default=-16.0)
args = p.parse_args()

TARGET = args.target

with open(args.data, encoding="utf-8") as f:
    records = json.load(f)

# sane vrijednosti — ebur128 vraća jako nizak I za tišinu
records = [r for r in records
           if isinstance(r.get("integrated_lufs"), (int, float))
           and r["integrated_lufs"] > -70]

# --- boje po |LUFS - target| ---
FILL_GREEN  = PatternFill("solid", fgColor="C6EFCE")  # ≤1 LU
FILL_YELLOW = PatternFill("solid", fgColor="FFEB9C")  # ≤2 LU
FILL_ORANGE = PatternFill("solid", fgColor="FCD5B4")  # ≤3 LU
FILL_RED    = PatternFill("solid", fgColor="FFC7CE")  # >3 LU
HDR_FILL    = PatternFill("solid", fgColor="1F3864")
HDR_FONT    = Font(bold=True, color="FFFFFF")
THIN        = Side(style="thin", color="D9D9D9")
BORDER      = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def dev_fill(lufs):
    d = abs(lufs - TARGET)
    if d <= 1:
        return FILL_GREEN
    if d <= 2:
        return FILL_YELLOW
    if d <= 3:
        return FILL_ORANGE
    return FILL_RED


def spread_fill(spread):
    if spread <= 3:
        return FILL_GREEN
    if spread <= 6:
        return FILL_YELLOW
    if spread <= 9:
        return FILL_ORANGE
    return FILL_RED


DATE_RE = re.compile(r"^(\d{8})_")
VID_RE = re.compile(r"_yt_([A-Za-z0-9_-]{11})(?:\.|$)")


def parse_base(base):
    date = ""
    m = DATE_RE.match(base)
    if m:
        d = m.group(1)
        date = f"{d[0:4]}-{d[4:6]}-{d[6:8]}"
    vid = ""
    mv = VID_RE.search(base)
    if mv:
        vid = mv.group(1)
    title = base
    title = DATE_RE.sub("", title)
    title = re.sub(r"_yt_[A-Za-z0-9_-]{11}.*$", "", title)
    title = title.replace("_", " ").strip()
    return date, title, vid


def style_header(ws, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(ncols)}{ws.max_row}"


wb = Workbook()

# ============ Sheet 1: Po kanalu ============
ws1 = wb.active
ws1.title = "Po kanalu"
hdr1 = ["Kanal", "N", "Median LUFS", "Mean", "Min", "Max", "Raspon (LU)", "Stddev", "Gain do cilja (med)"]
ws1.append(hdr1)

by_ch = {}
for r in records:
    by_ch.setdefault(r["channel"], []).append(r["integrated_lufs"])

ch_rows = []
for ch, arr in by_ch.items():
    arr_s = sorted(arr)
    med = statistics.median(arr_s)
    ch_rows.append({
        "ch": ch, "n": len(arr_s), "median": med,
        "mean": statistics.fmean(arr_s),
        "min": arr_s[0], "max": arr_s[-1],
        "spread": arr_s[-1] - arr_s[0],
        "stddev": statistics.pstdev(arr_s) if len(arr_s) > 1 else 0.0,
    })
ch_rows.sort(key=lambda x: x["median"])

for c in ch_rows:
    ws1.append([
        c["ch"], c["n"], round(c["median"], 2), round(c["mean"], 2),
        round(c["min"], 2), round(c["max"], 2), round(c["spread"], 2),
        round(c["stddev"], 2), round(TARGET - c["median"], 2),
    ])
    row = ws1.max_row
    ws1.cell(row=row, column=3).fill = dev_fill(c["median"])      # median
    ws1.cell(row=row, column=7).fill = spread_fill(c["spread"])   # raspon
    for col in range(1, len(hdr1) + 1):
        ws1.cell(row=row, column=col).border = BORDER

style_header(ws1, len(hdr1))
widths1 = [34, 6, 13, 9, 9, 9, 12, 9, 18]
for i, w in enumerate(widths1, 1):
    ws1.column_dimensions[get_column_letter(i)].width = w

# ============ Sheet 2: Epizode ============
ws2 = wb.create_sheet("Epizode")
hdr2 = ["Kanal", "Datum", "Naslov", "LUFS", "Odstupanje (vs cilj)", "Gain do cilja",
        "LRA (LU)", "Trajanje (min)", "Izvor", "Video ID"]
ws2.append(hdr2)

eps = sorted(records, key=lambda r: (r["channel"], r["integrated_lufs"]))
for r in eps:
    date, title, vid = parse_base(r["base"])
    lufs = r["integrated_lufs"]
    dur = r.get("duration_sec")
    ws2.append([
        r["channel"], date, title, round(lufs, 2), round(lufs - TARGET, 2),
        r.get("gain_to_target_db"),
        round(r["lra"], 2) if isinstance(r.get("lra"), (int, float)) else "",
        round(dur / 60, 1) if isinstance(dur, (int, float)) else "",
        r.get("source_type", ""), vid or r.get("video_id", "") or "",
    ])
    row = ws2.max_row
    f = dev_fill(lufs)
    ws2.cell(row=row, column=4).fill = f   # LUFS
    ws2.cell(row=row, column=5).fill = f   # Odstupanje
    for col in range(1, len(hdr2) + 1):
        ws2.cell(row=row, column=col).border = BORDER

style_header(ws2, len(hdr2))
widths2 = [30, 11, 48, 9, 18, 14, 10, 13, 8, 13]
for i, w in enumerate(widths2, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

# ============ Sheet 3: Sažetak ============
ws3 = wb.create_sheet("Sažetak")
all_lufs = sorted(r["integrated_lufs"] for r in records)
n = len(all_lufs)


def q(p):
    return all_lufs[min(n - 1, int(p * (n - 1)))]


within1 = sum(1 for v in all_lufs if abs(v - TARGET) <= 1)
within2 = sum(1 for v in all_lufs if abs(v - TARGET) <= 2)
off3 = sum(1 for v in all_lufs if abs(v - TARGET) > 3)

bold = Font(bold=True)
ws3["A1"] = "Glasnoća — sažetak"
ws3["A1"].font = Font(bold=True, size=14)
ws3["A2"] = f"Cilj: {TARGET} LUFS · epizoda: {n} · kanala: {len(by_ch)}"

summary = [
    ("", ""),
    ("Metrika", "Vrijednost"),
    ("Min", round(all_lufs[0], 2)),
    ("p10", round(q(0.10), 2)),
    ("Median", round(statistics.median(all_lufs), 2)),
    ("Mean", round(statistics.fmean(all_lufs), 2)),
    ("p90", round(q(0.90), 2)),
    ("Max", round(all_lufs[-1], 2)),
    ("RASPON (max-min)", round(all_lufs[-1] - all_lufs[0], 2)),
    ("Stddev", round(statistics.pstdev(all_lufs), 2)),
    ("", ""),
    (f"Unutar ±1 LU od cilja", f"{within1} ({within1/n*100:.0f}%)"),
    (f"Unutar ±2 LU od cilja", f"{within2} ({within2/n*100:.0f}%)"),
    (f">±3 LU od cilja", f"{off3} ({off3/n*100:.0f}%)"),
]
r0 = 4
for i, (k, v) in enumerate(summary):
    ws3.cell(row=r0 + i, column=1, value=k)
    ws3.cell(row=r0 + i, column=2, value=v)
    if k in ("Metrika",):
        ws3.cell(row=r0 + i, column=1).font = bold
        ws3.cell(row=r0 + i, column=2).font = bold

# histogram
hist_r = r0 + len(summary) + 1
ws3.cell(row=hist_r, column=1, value="Histogram (LUFS bucket → broj epizoda)").font = bold
lo, hi, step = -40, -5, 2.5
buckets = []
b = lo
while b < hi:
    buckets.append([b, b + step, 0])
    b += step
for v in all_lufs:
    idx = int((v - lo) / step)
    idx = max(0, min(len(buckets) - 1, idx))
    buckets[idx][2] += 1
hmax = max((bk[2] for bk in buckets), default=0)
hr = hist_r + 1
for lo_b, hi_b, cnt in buckets:
    if cnt == 0:
        continue
    ws3.cell(row=hr, column=1, value=f"{lo_b:.1f}..{hi_b:.1f}")
    ws3.cell(row=hr, column=2, value=cnt)
    bar = "█" * (round(cnt / hmax * 30) if hmax else 0)
    ws3.cell(row=hr, column=3, value=bar)
    hr += 1

note_r = hr + 1
note = ws3.cell(row=note_r, column=1, value=(
    "NAPOMENA: ebur128 mjeri 16kHz mono .wav → ~+3 LU offset vs full-band stereo izvora. "
    "RASPON između epizoda je valjan; apsolut nije (re-mjeri se u two-pass loudnorm-u). "
    "Boje: zeleno ≤1 LU od cilja, žuto ≤2, narančasto ≤3, crveno >3."))
note.alignment = Alignment(wrap_text=True)
ws3.merge_cells(start_row=note_r, start_column=1, end_row=note_r + 2, end_column=4)
ws3.column_dimensions["A"].width = 28
ws3.column_dimensions["B"].width = 16
ws3.column_dimensions["C"].width = 34

# legenda boja na Po kanalu sheetu (desno)
legend_col = len(hdr1) + 2
ws1.cell(row=1, column=legend_col, value="Legenda (|LUFS - cilj|)").font = bold
for i, (label, fill) in enumerate([
    ("≤1 LU", FILL_GREEN), ("≤2 LU", FILL_YELLOW),
    ("≤3 LU", FILL_ORANGE), (">3 LU", FILL_RED)], start=2):
    ws1.cell(row=i, column=legend_col, value=label).fill = fill
ws1.column_dimensions[get_column_letter(legend_col)].width = 22

# ============ Sheet 4: Rang kanala ============
# Objektivni rang: prava glasnoća (mjereno +OFFSET) blizu cilja + konzistentnost.
# Offset je sustavan (16kHz mono ebur128), pa je USPOREDBA među kanalima valjana.
OFFSET = 3.0
ws4 = wb.create_sheet("Rang kanala")
hdr4 = ["#", "Kanal", "N", "Median (mjereno)", "True median (+3)",
        "Δ od cilja", "Stddev", "Raspon (LU)", "Score (manje=bolje)"]
ws4.append(hdr4)

rank = []
for ch, arr in by_ch.items():
    if len(arr) < 3:
        continue
    a = sorted(arr)
    med = statistics.median(a)
    sd = statistics.pstdev(a) if len(a) > 1 else 0.0
    true_med = med + OFFSET
    dist = abs(true_med - TARGET)
    rank.append({
        "ch": ch, "n": len(a), "med": med, "true": true_med,
        "dist": dist, "sd": sd, "spread": a[-1] - a[0],
        "score": dist + 0.5 * sd,
    })
rank.sort(key=lambda x: x["score"])

for i, r in enumerate(rank, 1):
    ws4.append([
        i, r["ch"], r["n"], round(r["med"], 1), round(r["true"], 1),
        round(r["dist"], 1), round(r["sd"], 1), round(r["spread"], 1),
        round(r["score"], 1),
    ])
    row = ws4.max_row
    ws4.cell(row=row, column=6).fill = dev_fill(TARGET + r["dist"])   # Δ od cilja
    ws4.cell(row=row, column=8).fill = spread_fill(r["spread"])       # raspon
    # score: zeleno <2, žuto <4, narančasto <6, crveno ostalo
    sc = r["score"]
    ws4.cell(row=row, column=9).fill = (
        FILL_GREEN if sc < 2 else FILL_YELLOW if sc < 4
        else FILL_ORANGE if sc < 6 else FILL_RED)
    for col in range(1, len(hdr4) + 1):
        ws4.cell(row=row, column=col).border = BORDER

style_header(ws4, len(hdr4))
widths4 = [5, 30, 6, 17, 17, 11, 9, 12, 20]
for i, w in enumerate(widths4, 1):
    ws4.column_dimensions[get_column_letter(i)].width = w

wb.save(args.out)
print(f"Zapisano: {args.out} · epizoda: {n} · kanala: {len(by_ch)} · rangirano (n>=3): {len(rank)}")
